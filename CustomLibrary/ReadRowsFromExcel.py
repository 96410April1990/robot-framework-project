import openpyxl

def read_excel_rows(file, sheetName, columnNumber, rowNumber):
    count = 0
    wbObj = openpyxl.load_workbook(file)
    sheetObj = wbObj.get_sheet_by_name(sheetName)
    column = int(columnNumber)
    maxRow = sheetObj.max_row+1
    getRow = int(rowNumber)
    for i in range(getRow, maxRow):
        cellObj = sheetObj.cell(row=i, column = column)
        print(cellObj.value)
        count += 1
    return count